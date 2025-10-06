"""
Excel Exporter for Market Comparisons

Creates Excel files with multiple sheets (one per market pair).
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os


class MarketExcelExporter:
    """Exports market comparison data to Excel with multiple sheets"""
    
    def __init__(self, filename: str = None):
        """
        Initialize exporter
        
        Args:
            filename: Output filename (auto-generated if None)
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"market_comparisons_{timestamp}.xlsx"
        
        self.filename = filename
        self.sheets_data = {}
    
    def add_market_comparison(
        self,
        market_title: str,
        polymarket_data: Dict[str, Any],
        kalshi_data: Dict[str, Any],
        comparison: Dict[str, Any]
    ):
        """
        Add a market comparison to be exported
        
        Args:
            market_title: Title of the market
            polymarket_data: Polymarket market data
            kalshi_data: Kalshi market data
            comparison: Comparison results
        """
        # Create safe sheet name (Excel limit: 31 chars, no special chars)
        safe_title = market_title[:28].replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '')
        if safe_title in self.sheets_data:
            safe_title = f"{safe_title[:25]}_{len(self.sheets_data)}"
        
        # Build data rows
        data = []
        
        # Header
        data.append({
            'Metric': 'MARKET COMPARISON',
            'Polymarket': '',
            'Kalshi': '',
            'Difference': ''
        })
        
        data.append({
            'Metric': 'Market Title',
            'Polymarket': market_title,
            'Kalshi': market_title,
            'Difference': ''
        })
        
        data.append({
            'Metric': 'Market ID',
            'Polymarket': polymarket_data.get('market_id', ''),
            'Kalshi': kalshi_data.get('market_id', ''),
            'Difference': ''
        })
        
        # Separator
        data.append({'Metric': '', 'Polymarket': '', 'Kalshi': '', 'Difference': ''})
        
        # Prices
        data.append({
            'Metric': 'PRICES',
            'Polymarket': '',
            'Kalshi': '',
            'Difference': ''
        })
        
        poly_yes = None
        kalshi_yes = None
        
        if polymarket_data.get('outcomes'):
            for outcome in polymarket_data['outcomes']:
                if 'yes' in outcome.get('name', '').lower():
                    poly_yes = outcome.get('price', 0) * 100
                    data.append({
                        'Metric': f"  {outcome['name']} Price (%)",
                        'Polymarket': f"{poly_yes:.2f}%",
                        'Kalshi': '',
                        'Difference': ''
                    })
        
        if kalshi_data.get('outcomes'):
            for outcome in kalshi_data['outcomes']:
                if 'yes' in outcome.get('name', '').lower():
                    kalshi_yes = outcome.get('price', 0) * 100
                    data.append({
                        'Metric': f"  {outcome['name']} Price (%)",
                        'Polymarket': '',
                        'Kalshi': f"{kalshi_yes:.2f}%",
                        'Difference': ''
                    })
        
        if poly_yes and kalshi_yes:
            diff = abs(poly_yes - kalshi_yes)
            data[-1]['Difference'] = f"{diff:.2f}%"
        
        # Separator
        data.append({'Metric': '', 'Polymarket': '', 'Kalshi': '', 'Difference': ''})
        
        # Volume & Liquidity
        data.append({
            'Metric': 'VOLUME & LIQUIDITY',
            'Polymarket': '',
            'Kalshi': '',
            'Difference': ''
        })
        
        poly_vol = polymarket_data.get('total_volume', 0)
        kalshi_vol = kalshi_data.get('total_volume', 0)
        data.append({
            'Metric': '  Total Volume ($)',
            'Polymarket': f"${poly_vol:,.0f}" if poly_vol else 'N/A',
            'Kalshi': f"${kalshi_vol:,.0f}" if kalshi_vol else 'N/A',
            'Difference': f"${abs(poly_vol - kalshi_vol):,.0f}" if poly_vol and kalshi_vol else ''
        })
        
        poly_vol_24h = polymarket_data.get('raw_data', {}).get('volume_24hr', 0)
        kalshi_vol_24h = kalshi_data.get('raw_data', {}).get('volume_24h', 0)
        data.append({
            'Metric': '  Volume 24h ($)',
            'Polymarket': f"${poly_vol_24h:,.0f}" if poly_vol_24h else 'N/A',
            'Kalshi': f"${kalshi_vol_24h:,.0f}" if kalshi_vol_24h else 'N/A',
            'Difference': ''
        })
        
        poly_liq = polymarket_data.get('liquidity', 0)
        kalshi_liq = kalshi_data.get('liquidity', 0)
        data.append({
            'Metric': '  Liquidity ($)',
            'Polymarket': f"${poly_liq:,.0f}" if poly_liq else 'N/A',
            'Kalshi': f"${kalshi_liq:,.0f}" if kalshi_liq else 'N/A',
            'Difference': f"${abs(poly_liq - kalshi_liq):,.0f}" if poly_liq and kalshi_liq else ''
        })
        
        kalshi_oi = kalshi_data.get('raw_data', {}).get('open_interest', 0)
        data.append({
            'Metric': '  Open Interest (contracts)',
            'Polymarket': 'N/A',
            'Kalshi': f"{kalshi_oi:,}" if kalshi_oi else 'N/A',
            'Difference': ''
        })
        
        # Separator
        data.append({'Metric': '', 'Polymarket': '', 'Kalshi': '', 'Difference': ''})
        
        # Comparison Summary
        data.append({
            'Metric': 'COMPARISON SUMMARY',
            'Polymarket': '',
            'Kalshi': '',
            'Difference': ''
        })
        
        data.append({
            'Metric': '  Price Spread',
            'Polymarket': '',
            'Kalshi': '',
            'Difference': f"{comparison.get('price_spread', 0):.2f}%"
        })
        
        data.append({
            'Metric': '  Best Platform',
            'Polymarket': '✓' if comparison.get('best_platform', '').lower() == 'polymarket' else '',
            'Kalshi': '✓' if comparison.get('best_platform', '').lower() == 'kalshi' else '',
            'Difference': ''
        })
        
        data.append({
            'Metric': '  Arbitrage Opportunity',
            'Polymarket': '',
            'Kalshi': '',
            'Difference': 'YES' if comparison.get('arbitrage_opportunity') else 'NO'
        })
        
        # Separator
        data.append({'Metric': '', 'Polymarket': '', 'Kalshi': '', 'Difference': ''})
        
        # Timestamp
        data.append({
            'Metric': 'Timestamp',
            'Polymarket': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Kalshi': '',
            'Difference': ''
        })
        
        self.sheets_data[safe_title] = data
    
    def export(self) -> str:
        """
        Export all comparisons to Excel file
        
        Returns:
            Path to created file
        """
        if not self.sheets_data:
            print("⚠️  No data to export")
            return None
        
        # Create Excel writer
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            for sheet_name, data in self.sheets_data.items():
                # Create DataFrame
                df = pd.DataFrame(data)
                
                # Write to sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format the workbook
        self._format_workbook()
        
        print(f"✅ Exported to {self.filename}")
        print(f"   📊 {len(self.sheets_data)} market comparison sheets")
        
        return self.filename
    
    def _format_workbook(self):
        """Apply formatting to the workbook"""
        wb = load_workbook(self.filename)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Section headers (bold rows)
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            section_font = Font(bold=True)
            
            for row in ws.iter_rows(min_row=2):
                metric = row[0].value
                if metric and isinstance(metric, str) and metric.isupper():
                    for cell in row:
                        cell.fill = section_fill
                        cell.font = section_font
            
            # Column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
        
        wb.save(self.filename)


if __name__ == "__main__":
    # Test the exporter
    print("Testing Excel Exporter...\n")
    
    exporter = MarketExcelExporter("test_export.xlsx")
    
    # Test data
    poly_data = {
        "market_id": "0xtest123",
        "outcomes": [{"name": "Yes", "price": 0.054}],
        "total_volume": 2275,
        "liquidity": 1555,
        "raw_data": {"volume_24hr": 264}
    }
    
    kalshi_data = {
        "market_id": "KXTEST-26JAN01",
        "outcomes": [{"name": "Yes", "price": 0.02}],
        "total_volume": 5559,
        "liquidity": 2877,
        "raw_data": {"volume_24h": 70, "open_interest": 4375}
    }
    
    comparison = {
        "price_spread": 3.4,
        "best_platform": "polymarket",
        "arbitrage_opportunity": False
    }
    
    exporter.add_market_comparison(
        "Test Market - Will something happen?",
        poly_data,
        kalshi_data,
        comparison
    )
    
    exporter.export()
    
    # Cleanup
    if os.path.exists("test_export.xlsx"):
        os.remove("test_export.xlsx")
        print("\n✅ Test complete and cleaned up!")

